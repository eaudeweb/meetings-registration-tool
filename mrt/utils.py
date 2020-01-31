import os
import re
import urllib

import openpyxl
import collections
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from datetime import date, datetime
from json import JSONEncoder as _JSONEncoder
from PIL import Image
from unicodedata import normalize
from uuid import uuid4

from flask import _request_ctx_stack, current_app as app, g, url_for
from flask_babel import refresh
from flask_uploads import IMAGES, UploadSet

from werkzeug import FileStorage

from babel import support, Locale
from path import Path
from mrt.definitions import LANGUAGES_MAP, LANGUAGES_ISO_MAP

logos_upload = UploadSet('logos', IMAGES)


class Logo(object):
    def __init__(self, slug):
        self.default_filename = app.config[slug.upper()]

    @property
    def meeting_name(self):
        return 'meeting_{}_{}'.format(g.meeting.id, self.default_filename)

    @property
    def filename(self):
        if hasattr(g, 'meeting'):
            meeting_logo_name = self.meeting_name
            if logos_upload.path(meeting_logo_name).exists():
                return meeting_logo_name
        return self.default_filename

    @property
    def path(self):
        return logos_upload.path(self.filename)

    @property
    def url(self):
        if g.get('is_pdf_process'):
            return self.path
        filename = '{}/{}'.format(app.config['PATH_LOGOS_KEY'], self.filename)
        return url_for('files', filename=filename)

    @property
    def default(self):
        return self.default_filename == self.filename

    def save(self, data):
        self.unlink()
        logos_upload.save(data, name=self.meeting_name)

    def unlink(self):
        unlink_uploaded_file(self.meeting_name, 'logos')
        unlink_thumbnail_file(self.meeting_name, dir_name='logos')


def unlink_participant_custom_file(filename):
    unlink_uploaded_file(filename, 'custom')
    unlink_uploaded_file(filename, 'crop',
                         dir_name=app.config['PATH_CUSTOM_KEY'])
    unlink_thumbnail_file(filename, dir_name='custom_uploads')
    unlink_thumbnail_file(filename, dir_name='crops')


def unlink_uploaded_file(filename, config_key, dir_name=None):
    if filename:
        dir_path = app.config['UPLOADED_%s_DEST' % config_key.upper()]
        if dir_name:
            dir_path = dir_path / dir_name
        full_path = dir_path / filename
        if full_path.isfile():
            full_path.unlink()


def unlink_thumbnail_file(filename, dir_name=None):
    if filename:
        name, ext = Path(filename).splitext()
        thumb_path = app.config['UPLOADED_THUMBNAIL_DEST']
        if dir_name:
            thumb_path = thumb_path / dir_name
        if not thumb_path.exists():
            return
        for f in thumb_path.listdir():
            if f.basename().startswith(name):
                f.unlink()
        for dir_path in thumb_path.walkdirs():
            for f in dir_path.listdir():
                if f.basename().startswith(name):
                    f.unlink()


def duplicate_uploaded_file(filename, config_key):
    if filename:
        path_from_config = app.config['UPLOADED_%s_DEST' % config_key.upper()]
        full_path = path_from_config / filename
        if full_path.isfile():
            new_path = path_from_config / str(uuid4()) + full_path.ext
            full_path.copyfile(new_path)
            return new_path
    return False


def copy_attributes(destination, source, exclude_pk=True, exclude_fk=True,
                    exclude=[]):
    for col in destination.__table__.columns:
        if exclude_pk and col.primary_key:
            continue
        if exclude_fk and col.foreign_keys:
            continue
        if col.name in exclude:
            continue
        setattr(destination, col.name, getattr(source, col.name))
    return destination


_SLUG_RE = re.compile(r'[\t !"#$%&\'()*\-/<=>?@\[\\\]^_`{|},.]+')


def slugify(text, delim=u'-'):
    """Generates an slightly worse ASCII-only slug."""
    result = []
    for word in _SLUG_RE.split(text.lower()):
        word = normalize('NFKD', word).encode('ascii', 'ignore')
        if word:
            result.append(word)
    return unicode(delim.join(result))


def rotate_file(filename, config_key):
    """Rotates an image file, deletes the old one and returns the new filename
    """
    newfilename = str(uuid4()) + ".png"

    path_from_config = Path(
        app.config['UPLOADED_%s_DEST' % config_key.upper()]
    )

    try:
        img = Image.open(path_from_config / filename)
        img = img.rotate(-90)
        img.save(path_from_config / newfilename, "PNG")
    except IOError:
        return filename

    unlink_uploaded_file(filename, config_key)
    return newfilename


def crop_file(filename, config_key, data):
    path_from_config = Path(
        app.config['UPLOADED_%s_DEST' % config_key.upper()])
    img = Image.open(path_from_config / filename)
    img = img.crop(data)
    crop_path = (app.config['UPLOADED_CROP_DEST'] /
                 app.config['PATH_CUSTOM_KEY'])
    crop_path.makedirs_p()
    img.save(crop_path / filename)


def read_file(f):
    while True:
        data = f.read(131072)
        if data:
            yield data
        else:
            break
    f.close()


def generate_export_excel(header, rows, filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.row_dimensions[1].font = Font(bold=True)

    for col_idx, name in enumerate(header, 1):
        cell_col_letter = openpyxl.utils.get_column_letter(col_idx)
        sheet.cell(row=1, column=col_idx).value = name
        sheet.column_dimensions[cell_col_letter].width = len(name) + 1

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell_col_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.cell(row=row_idx, column=col_idx).value = unicode(value).encode('utf-8')

    workbook.save(filename)


def get_xlsx_header(fields):
    return collections.OrderedDict(
        (field.label.english + ' [required]' if field.required else field.label.english, field)
        for field in fields
    )


def generate_import_excel(fields, file_name, field_types, meeting_categories, countries):
    workbook = Workbook()
    sheet = workbook.active
    sheet.row_dimensions[1].font = Font(bold=True)
    col_names = get_xlsx_header(fields)

    val_sheet = workbook.create_sheet("Validation", index=1)
    val_sheet_col_idx = 1

    for col_idx, name in enumerate(col_names, 1):
        cell_col_letter = openpyxl.utils.get_column_letter(col_idx)

        sheet.cell(row=1, column=col_idx).value = name
        sheet.column_dimensions[cell_col_letter].width = len(name) + 1

        current_field = fields[col_idx - 1]
        if current_field.field_type.code == field_types.DATE:
            # Date may be bigger than column's header
            sheet.column_dimensions[cell_col_letter].width = max(len(name) + 1, 15)
            sheet.add_data_validation(
                DataValidation(
                    type="date",
                    error="The entry should be a date",
                    errorTitle="Invalid date",
                    sqref="{}2:{}2000".format(cell_col_letter, cell_col_letter),
                    operator="greaterThan",
                    formula1=date.min,
                    allow_blank=current_field.required,
                )
            )

        if (current_field.field_type.code == field_types.TEXT or\
                current_field.field_type.code == field_types.TEXT_AREA) and\
                current_field.max_length:

            sheet.add_data_validation(
                DataValidation(
                    type="textLength",
                    error="The entry can not be longer than {}".format(current_field.max_length),
                    errorTitle="Entry too long",
                    sqref="{}2:{}2000".format(cell_col_letter, cell_col_letter),
                    operator="lessThan",
                    formula1=current_field.max_length,
                    allow_blank=current_field.required,
                )
            )

        if current_field.field_type.code == field_types.MULTI_CHECKBOX:
            sheet.add_data_validation(
                DataValidation(
                    promptTitle = 'Fields selection',
                    prompt='Please type the entries comma separated (entry1, entry2, entry3)',
                    sqref="{}2:{}2000".format(cell_col_letter, cell_col_letter),
                    allow_blank=current_field.required,
                )
            )

        if current_field.field_type.code == field_types.SELECT or\
                current_field.field_type.code == field_types.RADIO or\
                current_field.field_type.code == field_types.LANGUAGE or\
                current_field.field_type.code == field_types.CHECKBOX or\
                current_field.field_type.code == field_types.CATEGORY or\
                current_field.field_type.code == field_types.COUNTRY:

            curr_validation_column = openpyxl.utils.get_column_letter(val_sheet_col_idx)
            cell = val_sheet.cell(1, val_sheet_col_idx, current_field.slug.upper())

            val_sheet.column_dimensions[curr_validation_column].width = 25

            values = [custom_val.value for custom_val in current_field.choices]
            if current_field.field_type.code == field_types.CHECKBOX:
                values = ["Yes", "No"]
            elif current_field.field_type.code == field_types.CATEGORY:
                values = meeting_categories
            elif current_field.field_type.code == field_types.COUNTRY:
                values = countries

            for row_index, value in enumerate(values, start=2):
                val_sheet.cell(row_index, val_sheet_col_idx, str(value))

            nr_values = len(values) + 1
            sheet.add_data_validation(
                DataValidation(
                    type="list",
                    error="Choose a value from the allowed entries",
                    errorTitle="Invalid entry",
                    formula1="'Validation'!${}$2:${}${}".format(curr_validation_column,
                                                                curr_validation_column,
                                                                nr_values),
                    sqref="{}2:{}2000".format(cell_col_letter, cell_col_letter),
                    allow_blank=current_field.required,
                )
            )

            val_sheet_col_idx += 1

    workbook.save(file_name)


def read_sheet(xlsx, fields, sheet_name=None):
    expected_headers = get_xlsx_header(fields)

    if sheet_name is None:
        sheet = xlsx.active
        sheet_name = sheet.title
    else:
        try:
            sheet = xlsx.get_sheet_by_name(sheet_name)
        except KeyError:
            raise ValueError("Missing sheet %r" % sheet_name)

    it = sheet.rows

    # Exclude empty cells.
    headers = [header.value.lower() for header in next(it) if header.value]
    # Lowercase the expected_headers
    expected_headers = {key.lower(): field for key, field in expected_headers.items()}
    # Check for consistency.
    difference = {h.lower() for h in expected_headers}.difference(set(headers))
    if difference:
        raise ValueError(
            "Missing column(s) %r in sheet %r" % (difference, sheet_name)
        )

    difference = set(headers).difference({h.lower() for h in expected_headers})
    if difference:
        raise ValueError(
            "Please remove column(s) %r in sheet %r" % (difference, sheet_name)
        )

    slug_headers = [expected_headers[header].slug for header in headers]
    # Iterate over the rows
    for row in it:
        row = [cell.value.strftime('%d.%m.%Y')
                if (isinstance(cell.value, date) or isinstance(cell.value, datetime))
                else (cell.value or "").encode('utf-8').decode('utf-8').strip() for cell in row[: len(headers)]]
        if not any(row):
            break
        yield dict(zip(slug_headers, row))


def get_translation(locale):
    ctx = _request_ctx_stack.top
    if ctx is None:
        return None
    translations = getattr(ctx, 'translation_%s' % locale.language, None)
    if translations is None:
        dirname = os.path.join(ctx.app.root_path, 'translations')
        translations = support.Translations.load(dirname, [locale])
        setattr(ctx, 'translation_%s' % locale.language, translations)
    return translations


def translate(text, lang_code='en'):
    locale = Locale(lang_code)
    translations = get_translation(locale)
    if translations:
        return translations.gettext(text).decode('unicode-escape')
    return text


def set_language(lang='english'):
    if lang in ('english', 'french', 'spanish'):
        iso = LANGUAGES_MAP.get(lang, 'en')
        g.language_verbose = lang
    if lang in ('en', 'fr', 'es'):
        iso = lang
        g.language_verbose = LANGUAGES_ISO_MAP.get(lang, 'english')
    g.language = iso
    refresh()


def validate_email(email):
    """Email validation function used by create_user command"""
    if re.match("^[a-zA-Z0-9._%\-+]+@[a-zA-Z0-9._%-]+.[a-zA-Z]{2,6}$", email):
        return True
    return False


def clean_email(emails):
    """Returns the first email from a string containing multiple emails"""
    for email in emails.split(','):
        email = email.strip()
        if validate_email(email):
            return email
    return None


def get_custom_file_as_filestorage(filename):
    file_path = app.config['UPLOADED_CUSTOM_DEST'] / filename
    try:
        data = FileStorage(stream=open(file_path, 'rb'),
                           filename=file_path.basename())
    except (IOError, OSError):
        data = None

    return data


class JSONEncoder(_JSONEncoder):

    def default(self, o):
        if isinstance(o, FileStorage):
            return str(o.filename)
        elif isinstance(o, date):
            return o.isoformat()
        return o


class CustomFieldLabel(object):

    def __init__(self, label):
        self.english = label.english
        self.french = label.french
        self.spanish = label.spanish

    def __unicode__(self):
        lang = getattr(g, 'language_verbose', 'english')
        return getattr(self, lang) or self.english


def parse_rfc6266_header(header_value):
    """Parse RFC6266 header and extract the filename."""
    header_value = header_value.strip()

    results = dict()
    for item in header_value.split(";"):
        item = item.strip()
        try:
            key, value = item.split("=")
        except ValueError:
            key, value = "", item

        charset = "ascii"
        encoding = ""
        if key.endswith("*"):
            # Possibly non-ascii value
            key = key.rstrip("*")
            # XXX Black magic afoot, as this is kinda standard but not quite.
            #  Although everybody implements it, but not quite.
            try:
                charset, encoding, value = value.split("'")
            except ValueError:
                try:
                    charset, value = value.split("'")
                except ValueError:
                    pass
        if encoding.lower() == "q":
            value = value.decode("quopri")
        elif encoding.lower() == "b":
            value = value.decode("base64")
        else:
            value = urllib.unquote(value)

        results[key] = value.decode(charset, "ignore").strip('"')
    return results

def str2bool(val, default_to_none=False):
    val = str(val).lower()
    if val in ("y", "yes", "t", "true", "on", "1"):
        return True
    elif val in ("n", "no", "f", "false", "off", "0"):
        return False
    elif val in ("null", "none"):
        return None
    elif not default_to_none:
        raise ValueError("invalid truth value %r" % (val,))
    else:
        return None
